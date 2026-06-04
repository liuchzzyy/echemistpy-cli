#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# ruff: noqa: N999
"""XLSX Data Reader for LANHE battery test files with metadata extraction using traitlets."""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, ClassVar

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd
import xarray as xr

from echemistpy.io.base_reader import BaseReader
from echemistpy.io.reader_utils import merge_infos, sanitize_variable_names
from echemistpy.io.structures import RawData, RawDataInfo

logger = logging.getLogger(__name__)


class LanheXLSXReader(BaseReader):
    """Reader for LANHE exported XLSX files.

    This reader handles the specific structure of LANHE battery test exports,
    including metadata from multiple sheets and time-series data from the main data sheet.
    """

    # --- Constants for better maintainability ---
    MEASUREMENT_COLUMNS: ClassVar[list[str]] = [
        "Record",
        "Cycle",
        "Step",
        "WorkMode",
        "StepInProcess",
        "StepDuration",
        "StepTime",
        "TestTime",
        "SysTime",
        "Voltage/V",
        "Current/uA",
        "Capacity/uAh",
        "SpeCap/mAh/g",
        "SpeCap_cal/mAh/g",
        "Energy/uWh",
        "SpeEnergy/mWh/g",
        "Power/uW",
        "dQdV/uAh/V",
        "dVdQ/V/uAh",
        "Temperature/C",
        "Temperature/℃",
        "Humidity/%",
        "Mark1",
        "Mark2",
        "BatteryCode",
        "DataFile",
        "TestName",
        "ProcessName",
        "Thicknessmm",
        "ThicknessPressureg",
        "ThicknessTempC",
        "ThicknessTemp℃",
        "ChannelNumber",
    ]

    METADATA_MAPPING: ClassVar[dict[str, str]] = {
        "Test name": "test_name",
        "Start time": "start_time",
        "Finish time": "finish_time",
        "Active material": "active_material",
        "Operator": "operator",
    }

    INDEX_COLUMNS: ClassVar[list[str]] = ["Record", "record", "Row", "row", "Index", "index"]
    TEXT_COLUMNS: ClassVar[set[str]] = {
        "BatteryCode",
        "Channel",
        "ChannelNumber",
        "DataFile",
        "Dev SN",
        "File name",
        "Log Details",
        "Log Type",
        "Mark1",
        "Mark2",
        "Path",
        "Process",
        "ProcessName",
        "Range",
        "Serial number",
        "StepInProcess",
        "Test name",
        "TestName",
        "WorkMode",
    }
    DATA_COLUMN_RENAMES: ClassVar[dict[str, str]] = {
        "StepDuration": "StepDuration_s",
        "StepTime": "StepTime_s",
        "TestTime": "TestTime_s",
        "Temperature/C": "Temperature_C",
        "Temperature/℃": "Temperature_C",
        "ThicknessTemp℃": "ThicknessTempC",
    }
    MASS_REGEX: ClassVar[re.Pattern[str]] = re.compile(r"([+-]?\d+(?:\.\d+)?)\s*(mg|g|ug|µg)", re.IGNORECASE)
    NUMERIC_REGEX: ClassVar[re.Pattern[str]] = re.compile(r"^[+-]?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?$")

    # --- Loader Metadata ---
    supports_directories: ClassVar[bool] = True
    instrument: ClassVar[str] = "lanhe"

    def __init__(self, filepath: str | Path | None = None, **kwargs: Any) -> None:
        """Initialize the LANHE reader.

        Args:
            filepath: Path to XLSX file or directory
            **kwargs: Additional metadata overrides
        """
        # Set default technique
        if "technique" not in kwargs:
            kwargs["technique"] = ["echem"]
        super().__init__(filepath, **kwargs)

    def _load_single_file(self, path: Path, **_kwargs: Any) -> tuple[RawData, RawDataInfo]:
        """Internal method to load and process a single LANHE XLSX file.

        Args:
            path: Path to the XLSX file
            **_kwargs: Additional arguments (unused, prefixed with _ to silence linter)

        Returns:
            Tuple of (RawData, RawDataInfo)
        """
        # 1. Extraction
        metadata, data_dict = self._read_xlsx(path)

        # 2. Cleaning & Standardization
        cleaned_metadata = self._clean_metadata(metadata)
        mass = self.active_material_mass or cleaned_metadata.get("active_material_mass")
        cleaned_data = self._clean_data(data_dict, active_material_mass=mass)

        # 3. Xarray Conversion
        ds = self._create_dataset(cleaned_data)
        ds = self._apply_time_coords(ds)
        ds = self._set_primary_index(ds)

        # 4. Metadata Packaging
        cleaned_metadata["file_path"] = str(path)
        if mass:
            cleaned_metadata["active_material_mass"] = mass
        raw_info = self._create_raw_info(cleaned_metadata, str(cleaned_metadata.get("test_name", path.stem)))

        return RawData(data=ds), raw_info

    @staticmethod
    def _create_dataset(data: dict[str, Any]) -> xr.Dataset:
        """Create an xarray Dataset from cleaned data dictionary.

        Args:
            data: Cleaned data dictionary

        Returns:
            xarray Dataset
        """
        data_vars = {LanheXLSXReader._dataset_column_name(k): (("record",), v) for k, v in data.items() if not k.startswith("_") and not LanheXLSXReader._all_missing(v)}
        if not data_vars:
            raise ValueError("No LANHE measurement records found in XLSX export.")

        ds = xr.Dataset(data_vars)
        ds = sanitize_variable_names(ds)
        if not isinstance(ds, xr.Dataset):
            raise TypeError("Expected sanitized LANHE data to remain an xarray.Dataset.")

        source_columns = data.get("_metadata", {}).get("columns", [])
        ds.attrs["source_columns"] = [str(c) for c in source_columns]
        return ds

    @staticmethod
    def _apply_time_coords(ds: xr.Dataset) -> xr.Dataset:
        """Convert SysTime to coordinates and calculate relative time.

        Args:
            ds: xarray Dataset

        Returns:
            Modified xarray Dataset
        """
        systime_key = "SysTime" if "SysTime" in ds else "SysTime".replace("/", "_")

        if systime_key in ds:
            ds = LanheXLSXReader._assign_systime_coord(ds, systime_key)

        test_time_key = "TestTime_s" if "TestTime_s" in ds else "TestTime"
        if test_time_key in ds:
            ds = LanheXLSXReader._assign_time_coord_from_column(ds, test_time_key)
        elif "systime" in ds.coords:
            systimes = pd.to_datetime(ds.coords["systime"].values)
            rel_times = (systimes - systimes[0]).total_seconds()
            ds = ds.assign_coords(time_s=(("record",), rel_times))
            ds.time_s.attrs.update({"units": "s", "long_name": "Relative Time"})

        return ds

    @staticmethod
    def _set_primary_index(ds: xr.Dataset) -> xr.Dataset:
        """Set the primary index for the 'record' dimension.

        Args:
            ds: xarray Dataset

        Returns:
            Modified xarray Dataset
        """
        for index_col in LanheXLSXReader.INDEX_COLUMNS:
            # Check both original and sanitized names
            for col in [index_col, index_col.replace("/", "_")]:
                if col in ds:
                    return ds.set_index(record=col)
        return ds

    def _create_raw_info(
        self,
        metadata: dict[str, Any],
        default_sample_name: str,
        technique_override: list[str] | None = None,
    ) -> RawDataInfo:
        """Create a RawDataInfo object from metadata.

        Args:
            path: File path
            metadata: Cleaned metadata dictionary
            mass: Active material mass

        Returns:
            RawDataInfo object
        """
        mass = metadata.get("active_material_mass")
        start_time_val = self.start_time or metadata.get("start_time")
        if isinstance(start_time_val, datetime):
            start_time_val = start_time_val.strftime("%Y-%m-%d %H:%M:%S")

        # Default to GCD if it's echem
        tech_list = list(self.technique)
        if tech_list == ["echem"]:
            tech_list.append("gcd")

        metadata_with_path = {**metadata}

        return RawDataInfo(
            sample_name=self.sample_name or default_sample_name,
            start_time=start_time_val,
            operator=self.operator or metadata.get("operator"),
            technique=technique_override or tech_list,
            instrument=self.instrument,
            active_material_mass=mass,
            wave_number=self.wave_number,
            others=metadata_with_path,
        )

    def _load_directory(self, path: Path, **_kwargs: Any) -> tuple[RawData, RawDataInfo]:
        """Load all LANHE XLSX files in a directory into a DataTree.

        Args:
            path: Path to the directory
            **_kwargs: Additional arguments (unused, prefixed with _ to silence linter)

        Returns:
            Tuple of (RawData with DataTree, merged RawDataInfo)
        """
        xlsx_files = sorted(path.rglob("*.xlsx"))
        if not xlsx_files:
            raise FileNotFoundError(f"No .xlsx files found in {path}")

        tree = xr.DataTree(name=path.name)
        infos = []

        for f in xlsx_files:
            try:
                raw_data, raw_info = self._load_single_file(f)
                node_path = "/".join(f.relative_to(path).with_suffix("").parts)

                tree[node_path] = raw_data.data
                tree[node_path].attrs.update(raw_info.to_dict())
                infos.append(raw_info)
            except Exception as e:
                logger.error("Failed to load %s: %s", f, e)

        if not tree.children and not tree.has_data:
            raise RuntimeError(f"Failed to load any valid .xlsx files from {path}")

        merged_info = merge_infos(
            infos,
            path,
            sample_name_override=self.sample_name,
            operator_override=self.operator,
            start_time_override=self.start_time,
            active_material_mass_override=self.active_material_mass,
            wave_number_override=self.wave_number,
            technique=list(self.technique),
            instrument=self.instrument,
        )

        return RawData(data=tree), merged_info

    def _read_xlsx(self, filepath: Path) -> tuple[dict[str, Any], dict[str, Any]]:
        """Read metadata and data from LANHE .xlsx file using openpyxl.

        Args:
            filepath: Path to the XLSX file

        Returns:
            Tuple of (metadata_dict, data_dict)
        """
        metadata: dict[str, Any] = {}
        data_dict: dict[str, Any] = {}

        # Use read_only for memory efficiency
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            metadata, data_dict = self._read_workbook(wb, filepath)
        finally:
            wb.close()

        return metadata, data_dict

    def _read_workbook(self, wb: openpyxl.Workbook, filepath: Path) -> tuple[dict[str, Any], dict[str, Any]]:
        """Read LANHE workbook metadata sheets and channel data sheet."""
        metadata: dict[str, Any] = {}
        self._read_test_info(wb, metadata)
        self._read_proc_info(wb, metadata)
        self._read_log_info(wb, metadata)

        data_sheet_name = self._find_data_sheet(wb)
        if data_sheet_name is None:
            logger.warning("No LANHE channel data sheet found in %s", filepath)
            return metadata, {}

        data_dict = self._read_record_data_from_ws(wb[data_sheet_name], data_sheet_name)
        sheet_metadata = data_dict.get("_metadata", {})
        if sheet_metadata:
            metadata["Channel_Data"] = sheet_metadata
        if cycle_summaries := sheet_metadata.get("cycle_summaries"):
            metadata["Cycle_Summary"] = cycle_summaries
        if step_summaries := sheet_metadata.get("step_summaries"):
            metadata["Step_Summary"] = step_summaries
        return metadata, data_dict

    def _read_record_data_from_ws(self, ws: openpyxl.worksheet.worksheet.Worksheet, sheet_name: str) -> dict[str, Any]:
        """Extract record-level data from an open worksheet.

        Args:
            ws: openpyxl Worksheet object
            sheet_name: Name of the sheet

        Returns:
            Dictionary with column data
        """
        header_pairs: list[tuple[int, str]] | None = None
        cycle_header_pairs: list[tuple[int, str]] | None = None
        step_header_pairs: list[tuple[int, str]] | None = None
        rows: list[tuple[Any, ...]] = []
        cycle_summaries: list[dict[str, Any]] = []
        step_summaries: list[dict[str, Any]] = []

        for row in ws.iter_rows(values_only=True):
            if self._is_empty_row(row):
                continue

            row_headers = self._header_values(row)
            if self._is_measurement_header(row_headers):
                header_pairs = self._header_pairs(row)
                continue
            if self._is_step_summary_header(row_headers):
                step_header_pairs = self._header_pairs(row)
                continue
            if self._is_cycle_summary_header(row_headers):
                cycle_header_pairs = self._header_pairs(row)
                continue

            if header_pairs and self._is_measurement_row(row):
                rows.append(row)
                continue
            if step_header_pairs and self._is_step_summary_row(row):
                step_summaries.append(self._row_to_dict(row, step_header_pairs))
                continue
            if cycle_header_pairs and self._is_cycle_summary_row(row):
                cycle_summaries.append(self._row_to_dict(row, cycle_header_pairs))
                continue

        if not (header_pairs and rows):
            return {}

        data = {header: [self._convert_cell_value(row[col_idx], header) if col_idx < len(row) else None for row in rows] for col_idx, header in header_pairs}
        data["_metadata"] = {
            "sheet_name": sheet_name,
            "num_rows": len(rows),
            "columns": [header for _, header in header_pairs],
            "cycle_summaries": cycle_summaries,
            "step_summaries": step_summaries,
        }
        return data

    def _read_test_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read 'Test information' sheet.

        Args:
            wb: openpyxl Workbook object
            metadata: Metadata dictionary to update
        """
        if "Test information" not in wb.sheetnames:
            return

        ws = wb["Test information"]
        headers = [str(cell.value).strip() for cell in ws[1] if cell.value]
        if ws.max_row >= 2:
            metadata["Test_Information"] = {h: self._convert_cell_value(ws[2][i].value, h) for i, h in enumerate(headers) if i < len(ws[2])}

    def _read_proc_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read 'Ch1_Proc' sheet and extract Work Mode table.

        Args:
            wb: openpyxl Workbook object
            metadata: Metadata dictionary to update
        """
        if "Ch1_Proc" not in wb.sheetnames:
            return
        ws = wb["Ch1_Proc"]
        proc_info: dict[str, Any] = {}
        work_mode: list[dict[str, Any]] = []
        headers: list[str] | None = None

        for row in ws.iter_rows(values_only=True):
            if self._is_empty_row(row) or row[0] is None:
                continue
            first_cell = str(row[0]).strip()
            if first_cell == "Order":
                headers = [str(c).strip() for c in row if c not in {None, ""}]
            elif headers:
                work_mode.append({h: self._convert_cell_value(row[i], h) for i, h in enumerate(headers) if i < len(row)})
            elif len(row) > 1 and row[1] is not None:
                proc_info[first_cell] = self._convert_cell_value(row[1], first_cell)

        if work_mode:
            proc_info["Work_Mode"] = work_mode
        metadata["Channel_Process_Info"] = proc_info

    def _read_log_info(self, wb: openpyxl.Workbook, metadata: dict[str, Any]) -> None:
        """Read 'Log' sheet.

        Args:
            wb: openpyxl Workbook object
            metadata: Metadata dictionary to update
        """
        if "Log" not in wb.sheetnames:
            return

        ws = wb["Log"]
        headers = [str(c.value).strip() for c in ws[1] if c.value]
        metadata["Log_Info"] = [
            {h: self._convert_cell_value(row[i], h) for i, h in enumerate(headers) if i < len(row)} for row in ws.iter_rows(min_row=2, values_only=True) if not self._is_empty_row(row)
        ]

    @staticmethod
    def _find_data_sheet(wb: openpyxl.Workbook) -> str | None:
        """Find the sheet containing 'DefaultGroup'.

        Args:
            wb: openpyxl Workbook object

        Returns:
            Sheet name or None
        """
        excluded = {"Test information", "Ch1_Proc", "Log"}
        return next((name for name in wb.sheetnames if "DefaultGroup" in name), None) or next((name for name in wb.sheetnames if name not in excluded), None)

    @staticmethod
    def _is_empty_row(row: tuple[Any, ...]) -> bool:
        """Return True when a worksheet row has no meaningful values."""
        return not row or all(cell is None or not str(cell).strip() for cell in row)

    @staticmethod
    def _header_values(row: tuple[Any, ...]) -> list[str]:
        """Extract non-empty string header values from a worksheet row."""
        return [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]

    @classmethod
    def _header_pairs(cls, row: tuple[Any, ...]) -> list[tuple[int, str]]:
        """Return unique non-empty headers with their source column indices."""
        seen: dict[str, int] = {}
        pairs: list[tuple[int, str]] = []
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            header = str(cell).strip()
            if not header:
                continue
            count = seen.get(header, 0)
            seen[header] = count + 1
            unique_header = header if count == 0 else f"{header}_{count + 1}"
            pairs.append((idx, unique_header))
        return pairs

    @staticmethod
    def _is_measurement_header(headers: list[str]) -> bool:
        """Detect the LANHE per-record measurement header."""
        return len(headers) >= 3 and headers[:3] == ["Cycle", "Step", "Record"]

    @staticmethod
    def _is_cycle_summary_header(headers: list[str]) -> bool:
        """Detect the cycle summary header at the top of the channel sheet."""
        return "CapC/uAh" in headers and "CapD/uAh" in headers

    @staticmethod
    def _is_step_summary_header(headers: list[str]) -> bool:
        """Detect a step summary header."""
        return len(headers) >= 2 and headers[0] == "Step" and headers[1] == "WorkMode"

    @staticmethod
    def _is_measurement_row(row: tuple[Any, ...]) -> bool:
        """LANHE record rows begin with numeric Cycle, Step, and Record values."""
        if len(row) < 3:
            return False
        try:
            int(row[0])
            int(row[1])
            int(row[2])
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _is_cycle_summary_row(row: tuple[Any, ...]) -> bool:
        """Cycle summary rows begin with a numeric cycle id."""
        try:
            int(row[0])
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _is_step_summary_row(row: tuple[Any, ...]) -> bool:
        """Step summary rows begin with numeric Step and textual WorkMode."""
        if len(row) < 2:
            return False
        try:
            int(row[0])
            return isinstance(row[1], str) and bool(row[1].strip())
        except (TypeError, ValueError):
            return False

    @classmethod
    def _row_to_dict(cls, row: tuple[Any, ...], header_pairs: list[tuple[int, str]]) -> dict[str, Any]:
        """Convert a worksheet row to a dictionary using source column indices."""
        return {header: cls._convert_cell_value(row[col_idx], header) if col_idx < len(row) else None for col_idx, header in header_pairs}

    @classmethod
    def _convert_cell_value(cls, value: Any, header: str | None = None) -> Any:
        """Convert one Excel cell into a Python value using LANHE conventions."""
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

        converted = cls._convert_time(value)
        if not isinstance(converted, str):
            return converted

        if header in cls.TEXT_COLUMNS:
            return converted
        if cls.NUMERIC_REGEX.match(converted):
            numeric = float(converted)
            return int(numeric) if numeric.is_integer() else numeric
        return converted

    @classmethod
    def _dataset_column_name(cls, column_name: str) -> str:
        """Return the xarray variable name for an official LANHE column name."""
        return cls.DATA_COLUMN_RENAMES.get(column_name, column_name)

    @staticmethod
    def _all_missing(values: Any) -> bool:
        """Return True when a column contains only missing values."""
        if not isinstance(values, list):
            return False
        return all(value is None for value in values)

    @staticmethod
    def _assign_systime_coord(ds: xr.Dataset, systime_key: str) -> xr.Dataset:
        """Assign absolute system time as a coordinate and remove duplicate variable."""
        try:
            systimes = pd.to_datetime(ds[systime_key].values)
        except (TypeError, ValueError) as exc:
            logger.warning("Failed to parse LANHE SysTime column: %s", exc)
            return ds

        ds = ds.assign_coords(systime=(("record",), systimes))
        ds = ds.drop_vars(systime_key)
        ds.systime.attrs.update({"long_name": "System Time"})
        return ds

    @staticmethod
    def _assign_time_coord_from_column(ds: xr.Dataset, time_key: str) -> xr.Dataset:
        """Assign relative test time as the canonical time_s coordinate."""
        try:
            time_values = pd.to_numeric(ds[time_key].values).astype(float)
        except (TypeError, ValueError) as exc:
            logger.warning("Failed to parse LANHE %s column as seconds: %s", time_key, exc)
            return ds

        ds = ds.assign_coords(time_s=(("record",), time_values))
        ds.time_s.attrs.update({"units": "s", "long_name": "Relative Test Time"})
        return ds

    @staticmethod
    def _convert_time(value: Any) -> Any:
        """Convert Excel values or LANHE time strings to standard formats.

        Args:
            value: Value to convert

        Returns:
            Converted value
        """
        if value is None or isinstance(value, (datetime, int, float)):
            return value

        if hasattr(value, "total_seconds"):
            return value.total_seconds()

        if not isinstance(value, str):
            return value

        return LanheXLSXReader._convert_time_string(value.strip())

    @staticmethod
    def _convert_time_string(value: str) -> Any:
        """Convert a LANHE time-like string to datetime, seconds, or original text."""
        if ":" not in value:
            return value
        days_duration = LanheXLSXReader._parse_days_duration(value)
        if days_duration is not None:
            return days_duration

        if " " in value:
            parts = value.split(" ", 1)
            parsed = LanheXLSXReader._parse_abs_time(parts[0], parts[1]) or LanheXLSXReader._parse_duration(parts[0], parts[1])
            return parsed if parsed is not None else value

        # Simple date YYYY-MM-DD
        if len(value) == 10 and value[4] in {"-", "/"}:
            try:
                return datetime.strptime(value.replace("/", "-"), "%Y-%m-%d")
            except ValueError:
                pass
        return value

    @staticmethod
    def _parse_days_duration(value: str) -> float | None:
        """Parse strings such as '0 days 06:43:26' into seconds."""
        match = re.match(r"^(\d+)\s+days?\s+(\d{1,2}:\d{2}:\d{2}(?:\.\d+)?)$", value)
        if not match:
            return None
        return LanheXLSXReader._parse_duration(match.group(1), match.group(2))

    @staticmethod
    def _parse_abs_time(date_part: str, time_part: str) -> datetime | None:
        """Parse YYYY-MM-DD HH:MM:SS.mmm format.

        Args:
            date_part: Date part of the string
            time_part: Time part of the string

        Returns:
            datetime object or None
        """
        try:
            # Normalize separator
            date_str = date_part.replace("/", "-")
            # Handle milliseconds if present
            if "." in time_part:
                return datetime.strptime(f"{date_str} {time_part}", "%Y-%m-%d %H:%M:%S.%f")
            return datetime.strptime(f"{date_str} {time_part}", "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None

    @staticmethod
    def _parse_duration(days_part: str, hms_part: str) -> float | None:
        """Parse D HH:MM:SS.mmm format into total seconds.

        Args:
            days_part: Days part of the string
            hms_part: Hours:Minutes:Seconds part

        Returns:
            Total seconds or None
        """
        if days_part.isdigit():
            try:
                hms = hms_part.split(":")
                if len(hms) == 3:
                    return int(days_part) * 86400 + int(hms[0]) * 3600 + int(hms[1]) * 60 + float(hms[2])
            except (ValueError, IndexError):
                pass
        return None

    def _clean_metadata(self, metadata: dict[str, Any]) -> dict[str, Any]:
        """Clean and map metadata to standard fields.

        Args:
            metadata: Raw metadata dictionary

        Returns:
            Cleaned metadata dictionary
        """
        cleaned: dict[str, Any] = {}

        if info := metadata.get("Test_Information"):
            for raw_key, clean_key in self.METADATA_MAPPING.items():
                actual_key = next((k for k in info if k.lower() == raw_key.lower()), None)
                if actual_key:
                    cleaned[clean_key] = info[actual_key]

            # Parse Active material mass and capacity
            if am := cleaned.get("active_material"):
                am_str = str(am)
                if "Active material:" in am_str:
                    parts = am_str.split("Active material:")
                    cleaned["active_material_mass"] = parts[1].strip()
                    if "Nominal specific capacity:" in parts[0]:
                        cleaned["nominal_specific_capacity"] = parts[0].replace("Nominal specific capacity:", "").strip()

        if proc := metadata.get("Channel_Process_Info"):
            proc_fields = ["Channel Number", "Name", "Description", "Unit Scheme", "Safety", "Work_Mode"]
            cleaned["channel_process_info"] = {k.lower().replace(" ", "_"): proc[k] for k in proc_fields if k in proc}

        cleaned["technique"] = list(self.technique)
        return {**metadata, **cleaned}

    def _clean_data(self, data: dict[str, Any], active_material_mass: Any = None) -> dict[str, Any]:
        """Filter, order, and optionally calculate specific capacity.

        Args:
            data: Raw data dictionary
            active_material_mass: Active material mass

        Returns:
            Cleaned data dictionary
        """
        # 1. Calculate Specific Capacity if mass is available
        spe_cap_cal = self._calculate_specific_capacity(data, active_material_mass)

        # 2. Build ordered result dictionary while preserving extra official columns
        cleaned_data = {}
        for col in self.MEASUREMENT_COLUMNS:
            if col == "SpeCap_cal/mAh/g":
                if spe_cap_cal is not None:
                    cleaned_data[col] = spe_cap_cal
            elif col in data:
                cleaned_data[col] = data[col]

        for col, values in data.items():
            if col.startswith("_") or col in cleaned_data:
                continue
            if not self._all_missing(values):
                cleaned_data[col] = values

        if "_metadata" in data:
            cleaned_data["_metadata"] = data["_metadata"]

        return cleaned_data

    @staticmethod
    def _calculate_specific_capacity(data: dict[str, Any], mass_input: Any) -> list[float | None] | None:
        """Calculate specific capacity (mAh/g) from capacity (uAh) and mass.

        Args:
            data: Data dictionary
            mass_input: Mass value

        Returns:
            List of specific capacity values or None
        """
        if not mass_input or "Capacity/uAh" not in data:
            return None

        mass_g = LanheXLSXReader._parse_mass_g(mass_input)
        if mass_g is None or mass_g <= 0:
            return None

        return LanheXLSXReader._specific_capacity_from_capacity(data["Capacity/uAh"], mass_g)

    @classmethod
    def _parse_mass_g(cls, mass_input: Any) -> float | None:
        """Parse active material mass into grams."""
        match = cls.MASS_REGEX.search(str(mass_input))
        if not match:
            return None

        value = float(match.group(1))
        unit = match.group(2).lower().replace("µ", "u")
        if unit == "g":
            return value
        if unit == "mg":
            return value * 1e-3
        if unit == "ug":
            return value * 1e-6
        return None

    @staticmethod
    def _specific_capacity_from_capacity(capacity_uah: list[Any], mass_g: float) -> list[float | None]:
        """Calculate specific capacity from capacity in uAh and mass in grams."""
        result: list[float | None] = []
        for capacity in capacity_uah:
            if capacity is None:
                result.append(None)
                continue
            try:
                result.append((float(capacity) / 1000.0) / mass_g)
            except (TypeError, ValueError, ZeroDivisionError):
                result.append(None)
        return result

    @staticmethod
    def _get_file_extension() -> str:
        """Get the file extension for this reader.

        Returns:
            File extension including the dot
        """
        return ".xlsx"
